"""
Facebook Auto-Poster — Google Drive + Excel Version
=====================================================
Excel থেকে schedule পড়ে, Google Drive থেকে ফাইল নামায়,
Facebook Page-এ পোস্ট করে।

ফোল্ডার structure (Google Drive):
  fb-autoposter/
  ├── posts/
  │   ├── images/   ← ছবি এখানে
  │   └── videos/   ← ভিডিও এখানে

Excel-এ ফাইলের নাম কলামে লিখবে:
  images/swert.jpg
  videos/sw1.mp4
"""

import os
import sys
import requests
import openpyxl
import tempfile
from pathlib import Path
from datetime import datetime

# ── Config ────────────────────────────────────────────
PAGE_ID       = os.environ.get("FACEBOOK_PAGE_ID",      "YOUR_PAGE_ID")
ACCESS_TOKEN  = os.environ.get("FACEBOOK_ACCESS_TOKEN", "YOUR_TOKEN")
GDRIVE_FOLDER = os.environ.get("GDRIVE_FOLDER_ID",      "1HpyhZxShboI7vfMT7NmviAueUhddmU3m")
EXCEL_FILE    = Path("facebook_content_calendar.xlsx")
SHEET_NAME    = "Content Calendar"

# Excel column index (1-based)
COL_ID        = 1
COL_FILENAME  = 2
COL_TYPE      = 3
COL_CAPTION   = 4
COL_SCHEDULE  = 5
COL_STATUS    = 6
COL_POST_ID   = 7
COL_NOTE      = 8

BASE_URL   = f"https://graph.facebook.com/v19.0/{PAGE_ID}"
DRIVE_API  = "https://www.googleapis.com/drive/v3"
DRIVE_DOWN = "https://drive.google.com/uc?export=download"

# ── Google Drive helpers ──────────────────────────────

def get_file_id(filename):
    """
    filename = "images/swert.jpg"
    → Google Drive-এ fb-autoposter/images/swert.jpg খোঁজো
    → file ID return করো
    """
    parts = filename.strip("/").split("/")

    # প্রথমে subfolder খোঁজো (images বা videos)
    subfolder_name = parts[0]   # e.g. "images"
    file_name      = parts[1]   # e.g. "swert.jpg"

    # Subfolder ID খোঁজো
    url = f"{DRIVE_API}/files"
    params = {
        "q": f"'{GDRIVE_FOLDER}' in parents and name='{subfolder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false",
        "fields": "files(id,name)",
        "supportsAllDrives": "true",
    }
    resp = requests.get(url, params=params).json()
    folders = resp.get("files", [])

    if not folders:
        print(f"  ❌ '{subfolder_name}' subfolder পাওয়া যায়নি Google Drive-এ")
        return None

    subfolder_id = folders[0]["id"]

    # File ID খোঁজো
    params2 = {
        "q": f"'{subfolder_id}' in parents and name='{file_name}' and trashed=false",
        "fields": "files(id,name)",
        "supportsAllDrives": "true",
    }
    resp2 = requests.get(url, params=params2).json()
    files = resp2.get("files", [])

    if not files:
        print(f"  ❌ '{file_name}' ফাইল পাওয়া যায়নি Google Drive-এ")
        return None

    return files[0]["id"]

def download_file(file_id, dest_path):
    """Google Drive থেকে ফাইল download করো।"""
    # প্রথম request
    session = requests.Session()
    resp = session.get(DRIVE_DOWN, params={"id": file_id}, stream=True)

    # Large file confirmation (virus scan warning bypass)
    token = None
    for key, value in resp.cookies.items():
        if key.startswith("download_warning"):
            token = value
            break

    if token:
        resp = session.get(
            DRIVE_DOWN,
            params={"id": file_id, "confirm": token},
            stream=True
        )

    with open(dest_path, "wb") as f:
        for chunk in resp.iter_content(chunk_size=32768):
            if chunk:
                f.write(chunk)

    size = Path(dest_path).stat().st_size
    print(f"  ✓ Download হয়েছে: {Path(dest_path).name} ({size//1024} KB)")
    return size > 0

# ── Load Excel ────────────────────────────────────────

def load_sheet():
    if not EXCEL_FILE.exists():
        print(f"❌ Excel file পাওয়া যায়নি: {EXCEL_FILE}")
        sys.exit(1)
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    return wb, ws

def save_sheet(wb):
    wb.save(EXCEL_FILE)

def get_due_rows(ws):
    now = datetime.now()
    due = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        status   = row[COL_STATUS - 1].value
        schedule = row[COL_SCHEDULE - 1].value
        if status and str(status).strip().lower() == "pending":
            if schedule and isinstance(schedule, datetime) and schedule <= now:
                due.append(row)
    return due

# ── Post functions ────────────────────────────────────

def post_image(file_path, caption):
    print(f"  📸 Image পোস্ট করছি...")
    with open(file_path, "rb") as f:
        resp = requests.post(
            f"{BASE_URL}/photos",
            params={"access_token": ACCESS_TOKEN},
            files={"source": f},
            data={"caption": caption or "", "published": "true"}
        ).json()
    if "id" in resp:
        return resp["id"]
    print(f"  ❌ Image error: {resp.get('error', {}).get('message', resp)}")
    return None

def post_video(file_path, caption):
    print(f"  🎬 Video পোস্ট করছি...")
    file_size = Path(file_path).stat().st_size

    # Init
    init = requests.post(
        f"{BASE_URL}/video_reels",
        data={"upload_phase": "start", "access_token": ACCESS_TOKEN}
    ).json()
    if "error" in init:
        print(f"  ❌ Video init error: {init['error']['message']}")
        return None

    video_id   = init["video_id"]
    upload_url = init["upload_url"]

    # Upload
    with open(file_path, "rb") as f:
        up = requests.post(
            upload_url,
            headers={
                "Authorization": f"OAuth {ACCESS_TOKEN}",
                "offset": "0",
                "file_size": str(file_size),
            },
            data=f
        ).json()
    if not up.get("success"):
        print(f"  ❌ Upload error: {up}")
        return None

    # Publish
    pub = requests.post(
        f"{BASE_URL}/video_reels",
        data={
            "upload_phase": "finish",
            "access_token": ACCESS_TOKEN,
            "video_id": video_id,
            "video_state": "PUBLISHED",
            "description": caption or "",
        }
    ).json()
    if "error" in pub:
        print(f"  ❌ Publish error: {pub['error']['message']}")
        return None
    return video_id

def post_text(caption):
    print(f"  📝 Text পোস্ট করছি...")
    resp = requests.post(
        f"{BASE_URL}/feed",
        data={"message": caption or "", "access_token": ACCESS_TOKEN}
    ).json()
    if "id" in resp:
        return resp["id"]
    print(f"  ❌ Text error: {resp.get('error', {}).get('message', resp)}")
    return None

def mark_done(ws, row, post_id):
    from openpyxl.styles import PatternFill, Font
    row[COL_STATUS - 1].value  = "Done"
    row[COL_STATUS - 1].fill   = PatternFill("solid", fgColor="D4EDDA")
    row[COL_STATUS - 1].font   = Font(name="Arial", size=10, bold=True, color="155724")
    row[COL_POST_ID - 1].value = str(post_id)

# ── Main ──────────────────────────────────────────────

def main():
    print("=" * 55)
    print("  Facebook Auto-Poster (Google Drive + Excel)")
    print(f"  সময়: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("=" * 55)

    wb, ws = load_sheet()
    due_rows = get_due_rows(ws)

    if not due_rows:
        print("ℹ️  এই মুহূর্তে কোনো scheduled পোস্ট নেই।")
        return

    print(f"\n📋 {len(due_rows)}টি পোস্ট পাওয়া গেছে।\n")
    posted_count = 0

    for row in due_rows:
        row_num   = row[0].row
        filename  = str(row[COL_FILENAME - 1].value or "").strip()
        post_type = str(row[COL_TYPE - 1].value or "").strip().lower()
        caption   = str(row[COL_CAPTION - 1].value or "").strip()

        print(f"── Row {row_num}: {filename or '(text only)'} [{post_type}]")

        post_id = None

        if post_type == "text" or not filename:
            post_id = post_text(caption)

        elif post_type in ("image", "video"):
            # Google Drive থেকে ফাইল নামাও
            print(f"  ☁️  Google Drive থেকে নামাচ্ছি: {filename}")
            file_id = get_file_id(filename)
            if not file_id:
                continue

            # Temp ফাইলে save করো
            suffix = Path(filename).suffix
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                tmp_path = tmp.name

            if not download_file(file_id, tmp_path):
                print(f"  ❌ Download ব্যর্থ")
                continue

            if post_type == "image":
                post_id = post_image(tmp_path, caption)
            else:
                post_id = post_video(tmp_path, caption)

            # Temp ফাইল মুছে দাও
            Path(tmp_path).unlink(missing_ok=True)

        else:
            print(f"  ⚠️  অচেনা ধরন: '{post_type}'")
            continue

        if post_id:
            mark_done(ws, row, post_id)
            print(f"  ✅ সফল! Post ID: {post_id}")
            posted_count += 1
        else:
            print(f"  ❌ ব্যর্থ।")

    save_sheet(wb)
    print(f"\n{'='*55}")
    print(f"  ✅ {posted_count}/{len(due_rows)} পোস্ট সফল।")
    print(f"  Excel আপডেট হয়েছে।")
    print(f"{'='*55}")

if __name__ == "__main__":
    main()
